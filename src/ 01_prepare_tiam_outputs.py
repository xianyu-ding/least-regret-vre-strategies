#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep  1 17:26:31 2025

@author: vanessa
"""

import pandas as pd
import numpy as np
import os
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def generate_tiam_new_installed_excel():
    """
    Generate an Excel file with NEW INSTALLED capacity data based on 
    TIAM model outputs for two periods: 2020-2030 and 2030-2050.
    This calculates incremental capacity additions rather than total capacity.
    """
    # Define the input file paths
    cap_data_path = '/Users/vanessa/Library/CloudStorage/OneDrive-UniversityCollegeLondon/research/data/exported/042325_230617863.csv'
    gen_data_path = '/Users/vanessa/Library/CloudStorage/OneDrive-UniversityCollegeLondon/research/data/exported/042325_230457106.csv'
    
    # Define output paths - save to the same location as script 1 expects
    output_path = '/Users/vanessa/Library/CloudStorage/OneDrive-UniversityCollegeLondon/research/data/paper2/TIAM_data_output_newinstalled.xlsx'
    matrix_output_path = '/Users/vanessa/Library/CloudStorage/OneDrive-UniversityCollegeLondon/research/data/paper2/TIAM_matrix_newinstalled.xlsx'
    
    print(f"Loading capacity data from: {cap_data_path}")
    cap_data = pd.read_csv(cap_data_path)
    
    print(f"Loading generation data from: {gen_data_path}")
    gen_data = pd.read_csv(gen_data_path)
    
    # Convert PJ to GWh for generation data
    gen_data['Pv'] = gen_data['Pv'] * 277.778  # 1 PJ = 277.778 GWh
    
    # Only exclude onshore and offshore wind
    exclude_sets = ['ELC FROM WIND-ONSH', 'ELC FROM WIND-OFFSH']
    
    cap_data = cap_data[~cap_data['Processset'].isin(exclude_sets)].copy()
    gen_data = gen_data[~gen_data['Processset'].isin(exclude_sets)].copy()
    
    # Focus on periods: 2020, 2030, 2050
    periods_of_interest = [2020, 2030, 2050]
    cap_data = cap_data[cap_data['Period'].isin(periods_of_interest)].copy()
    gen_data = gen_data[gen_data['Period'].isin(periods_of_interest)].copy()
    
    # Find all "net-xx" scenarios in the data
    net_scenarios = [s for s in cap_data['Scenario'].unique() if s.startswith('netzero-') or s == 'ndc-test3-netzero']
    
    # Keep only net scenarios and ndc-test3 as baseline
    scenarios = net_scenarios + ['ndc-test3']
    filtered_cap_data = cap_data[cap_data['Scenario'].isin(scenarios)].copy()
    filtered_gen_data = gen_data[gen_data['Scenario'].isin(scenarios)].copy()
    
    # Separate data for special processing
    biomass_cap_data = filtered_cap_data[filtered_cap_data['Processset'] == 'ELC FROM BIOMASS'].copy()
    bio_ccs_cap_data = filtered_cap_data[filtered_cap_data['Processset'] == 'ELC FROM BIO CCS'].copy()
    gas_ccs_cap_data = filtered_cap_data[filtered_cap_data['Processset'] == 'ELC FROM GAS CCS'].copy()
    coal_ccs_cap_data = filtered_cap_data[filtered_cap_data['Processset'] == 'ELC FROM COAL CCS'].copy()
    
    # Set regular technology mapping (now with separate CCS technologies)
    name_mapping = {
        'ELC FROM HYDRO': 'Hydro',
        'ELC FROM WIND': 'Wind',
        'ELC FROM NUCLEAR': 'Nuclear',
        'ELC FROM OIL': 'Oil',
        'ELC FROM GAS': 'Gas',
        'ELC FROM COALS': 'Coal',
        'ELC FROM SOL-THERM': 'Solar',
        'ELC FROM SOL-PV': 'Solar',
        'ELC FROM GEO': 'Geo',
        'ELC FROM STG': 'Storage',
        'ELC FROM TIDAL': 'Tidal',
        # Map to special categories for separate processing
        'ELC FROM BIOMASS': 'Biomass_Special',
        'ELC FROM BIO CCS': 'Bio_CCS_Special',
        'ELC FROM GAS CCS': 'Gas_CCS_Special',
        'ELC FROM COAL CCS': 'Coal_CCS_Special',
    }
    
    filtered_cap_data.loc[:, 'Tech'] = filtered_cap_data['Processset'].map(name_mapping).fillna(filtered_cap_data['Processset'])
    filtered_gen_data.loc[:, 'Tech'] = filtered_gen_data['Processset'].map(name_mapping).fillna(filtered_gen_data['Processset'])
    
    # Add region classification
    def classify_region(process):
        suffix_to_region = {
            '15': 'NWR', 'GEN01': 'NWR',
            '25': 'NR', 'GEN02': 'NR',
            '35': 'NER', 'GEN03': 'NER',
            '45': 'ER', 'GEN04': 'ER',
            '55': 'CR', 'GEN05': 'CR',
            '65': 'SWR', 'GEN06': 'SWR',
            '75': 'SR', 'GEN07': 'SR',
            'ESTGHYDPMP1': 'NWR', 'ESTGHYDPMP2': 'NR', 'ESTGHYDPMP3': 'NER',
            'ESTGHYDPMP4': 'ER', 'ESTGHYDPMP5': 'CR', 'ESTGHYDPMP6': 'SWR', 'ESTGHYDPMP7': 'SR',
            'ESTGBESSC1': 'NWR', 'ESTGBESSC2': 'NR', 'ESTGBESSC3': 'NER',
            'ESTGBESSC4': 'ER', 'ESTGBESSC5': 'CR', 'ESTGBESSC6': 'SWR', 'ESTGBESSC7': 'SR',
            'ESTGBESSD1': 'NWR', 'ESTGBESSD2': 'NR', 'ESTGBESSD3': 'NER',
            'ESTGBESSD4': 'ER', 'ESTGBESSD5': 'CR', 'ESTGBESSD6': 'SWR', 'ESTGBESSD7': 'SR'
        }
        for suffix, region in suffix_to_region.items():
            if process.endswith(suffix):
                return region 
        return None
    
    filtered_cap_data.loc[:, 'RegionClass'] = filtered_cap_data['Process'].apply(classify_region)
    filtered_gen_data.loc[:, 'RegionClass'] = filtered_gen_data['Process'].apply(classify_region)
    
    # Make sure those interpolated entries are still labeled as 'Storage'
    filtered_cap_data.loc[filtered_cap_data['Process'].str.contains('BESS|HYDPMP', na=False), 'Tech'] = 'Storage'
    filtered_gen_data.loc[filtered_gen_data['Process'].str.contains('BESS|HYDPMP', na=False), 'Tech'] = 'Storage'
    
    # Define the ordered regions and technologies
    regions = ['NWR', 'NR', 'NER', 'ER', 'CR', 'SWR', 'SR']
    all_techs = ['Coal', 'Gas', 'Geo', 'Hydro', 'Nuclear', 'Oil', 'Solar', 'Storage', 'Tidal', 'Wind', 'Biomass', 
                'Bio_CCS', 'Gas_CCS', 'Coal_CCS']
    
    # Process scenarios in a specific order
    scenario_order = ['ndc-test3-netzero']  # Start with net-zero
    # Add other net scenarios in descending order (90, 85, 80, etc.)
    for i in range(95, 50, -5):
        scenario = f'netzero-pr{i}-2030cap'
        if scenario in scenarios:
            scenario_order.append(scenario)
    scenario_order.append('ndc-test3')  # Add ndc-test3 at the end
    
    # Create a mapping for scenario names to sheet names - BUT we need TWO periods
    scenario_names = {
        'ndc-test3': 'ndc',
        'ndc-test3-netzero': 'net'
    }
    # Add net-xx scenarios
    for i in range(5, 100, 5):
        scenario = f'netzero-pr{i}-2030cap'
        scenario_names[scenario] = f'net-{i}'
    
    def calculate_new_installed_capacity(scenario_data, biomass_data, bio_ccs_data, gas_ccs_data, coal_ccs_data, period_start, period_end):
        """
        Calculate new installed capacity for a given period by subtracting start capacity from end capacity
        """
        # Get capacity data for both periods
        start_data = scenario_data[scenario_data['Period'] == period_start]
        end_data = scenario_data[scenario_data['Period'] == period_end]
        
        # Aggregate by region and technology for both periods
        start_cap = start_data.groupby(['RegionClass', 'Tech'])['Pv'].sum().unstack(fill_value=0)
        end_cap = end_data.groupby(['RegionClass', 'Tech'])['Pv'].sum().unstack(fill_value=0)
        
        # Ensure both dataframes have the same structure
        all_regions_techs = set(start_cap.index.tolist() + end_cap.index.tolist())
        all_tech_cols = set(start_cap.columns.tolist() + end_cap.columns.tolist())
        
        start_cap = start_cap.reindex(index=all_regions_techs, columns=all_tech_cols, fill_value=0)
        end_cap = end_cap.reindex(index=all_regions_techs, columns=all_tech_cols, fill_value=0)
        
        # Calculate new installations (end - start)
        new_installed = end_cap - start_cap
        
        # Handle negative values (capacity retirement) - set to 0 for environmental impact calculation
        new_installed = new_installed.clip(lower=0)
        
        # Process special technologies (Biomass, Bio CCS, Gas CCS, Coal CCS)
        for special_data, target_col in [(biomass_data, 'Biomass'), (bio_ccs_data, 'Bio_CCS'), 
                                       (gas_ccs_data, 'Gas_CCS'), (coal_ccs_data, 'Coal_CCS')]:
            start_special = special_data[special_data['Period'] == period_start]['Pv'].sum()
            end_special = special_data[special_data['Period'] == period_end]['Pv'].sum()
            
            # Calculate new installation for this technology
            new_special = max(0, end_special - start_special)  # Clip negative to 0
            new_special_per_region = new_special / 7  # Distribute equally across 7 regions
            
            # Ensure column exists
            if target_col not in new_installed.columns:
                new_installed[target_col] = 0
            
            # Add the new installation to each region
            for region in regions:
                if region in new_installed.index:
                    new_installed.loc[region, target_col] = new_special_per_region
        
        # Remove special technology columns if they exist
        special_techs = ['Biomass_Special', 'Bio_CCS_Special', 'Gas_CCS_Special', 'Coal_CCS_Special']
        for special_tech in special_techs:
            if special_tech in new_installed.columns:
                new_installed = new_installed.drop(columns=[special_tech])
        
        # Ensure all technology columns exist
        for tech in all_techs:
            if tech not in new_installed.columns:
                new_installed[tech] = 0
                
        # Reorder columns to match the desired order
        new_installed = new_installed[all_techs]
        
        # Reindex to ensure all regions are included in correct order
        new_installed = new_installed.reindex(regions, fill_value=0)
        
        return new_installed
    
    # Create matrix DataFrames for data checking - we'll create separate files for each period
    matrix_dfs_2030 = {}  # 2020-2030 period
    matrix_dfs_2050 = {}  # 2030-2050 period
    
    # Create Excel writers for both periods
    output_path_2030 = output_path.replace('.xlsx', '_2020_2030.xlsx')
    output_path_2050 = output_path.replace('.xlsx', '_2030_2050.xlsx')
    
    # Process 2020-2030 period
    print(f"Creating Excel file for 2020-2030 period: {output_path_2030}")
    with pd.ExcelWriter(output_path_2030, engine='openpyxl') as writer:
        
        for scenario in scenario_order:
            if scenario not in scenarios:
                continue
                
            sheet_name = scenario_names.get(scenario, scenario[:10])
            print(f"Processing scenario: {scenario} (Sheet: {sheet_name}) for 2020-2030")
            
            # Extract capacity data for this scenario
            scenario_cap_data = filtered_cap_data[filtered_cap_data['Scenario'] == scenario]
            
            # Extract special technology data
            biomass_scenario = biomass_cap_data[biomass_cap_data['Scenario'] == scenario]
            bio_ccs_scenario = bio_ccs_cap_data[bio_ccs_cap_data['Scenario'] == scenario]
            gas_ccs_scenario = gas_ccs_cap_data[gas_ccs_cap_data['Scenario'] == scenario]
            coal_ccs_scenario = coal_ccs_cap_data[coal_ccs_cap_data['Scenario'] == scenario]
            
            # Skip if no data for this scenario
            if len(scenario_cap_data) == 0:
                print(f"No data for scenario: {scenario}, skipping...")
                continue
            
            # Calculate new installed capacity for 2020-2030
            new_installed_cap = calculate_new_installed_capacity(
                scenario_cap_data, biomass_scenario, bio_ccs_scenario, 
                gas_ccs_scenario, coal_ccs_scenario, 2020, 2030
            )
            
            # Store the data for the matrix DataFrame
            matrix_dfs_2030[scenario] = new_installed_cap.copy()
            
            # Process hydro generation data for 2030
            hydro_gen_2030 = filtered_gen_data[(filtered_gen_data['Scenario'] == scenario) & 
                                             (filtered_gen_data['Tech'] == 'Hydro') & 
                                             (filtered_gen_data['Period'] == 2030)]
            hydro_act_series = hydro_gen_2030.groupby('RegionClass')['Pv'].sum()
            hydro_act = pd.Series(0, index=new_installed_cap.index)
            for region in hydro_act_series.index:
                if region in hydro_act.index:
                    hydro_act[region] = hydro_act_series[region]
            
            # Calculate additional columns
            new_installed_cap['Sum'] = new_installed_cap.drop(columns=['Storage'], errors='ignore').sum(axis=1)
            
            with np.errstate(divide='ignore', invalid='ignore'):
                new_installed_cap['Storage_Ratio'] = new_installed_cap['Storage'] / new_installed_cap['Sum']
            
            new_installed_cap['Wind_Solar_Sum'] = new_installed_cap['Wind'] + new_installed_cap['Solar']
            
            with np.errstate(divide='ignore', invalid='ignore'):
                new_installed_cap['Wind_Solar_Ratio'] = new_installed_cap['Wind_Solar_Sum'] / new_installed_cap['Sum']
            
            new_installed_cap = new_installed_cap.fillna(0)
            new_installed_cap['Hydro_act'] = hydro_act
            
            # Calculate total row
            total_series = {}
            for col in new_installed_cap.columns:
                if col not in ['Storage_Ratio', 'Wind_Solar_Ratio']:
                    total_series[col] = new_installed_cap[col].sum()
            
            with np.errstate(divide='ignore', invalid='ignore'):
                total_series['Storage_Ratio'] = total_series.get('Storage', 0) / total_series.get('Sum', 1) if total_series.get('Sum', 0) != 0 else 0
                total_series['Wind_Solar_Ratio'] = total_series.get('Wind_Solar_Sum', 0) / total_series.get('Sum', 1) if total_series.get('Sum', 0) != 0 else 0
            
            # Convert to DataFrame for Excel output
            new_installed_cap_df = new_installed_cap.reset_index()
            new_installed_cap_df.rename(columns={'index': 'RegionClass'}, inplace=True)
            
            # Append total row
            total_df = pd.DataFrame([['Total'] + [total_series.get(col, 0) for col in new_installed_cap.columns]], 
                                   columns=['RegionClass'] + list(new_installed_cap.columns))
            new_installed_cap_df = pd.concat([new_installed_cap_df, total_df], ignore_index=True)
            new_installed_cap_df = new_installed_cap_df.fillna(0)
            
            # Write to Excel
            new_installed_cap_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the worksheet
            format_worksheet(writer.sheets[sheet_name], new_installed_cap_df)
    
    # Process 2030-2050 period
    print(f"Creating Excel file for 2030-2050 period: {output_path_2050}")
    with pd.ExcelWriter(output_path_2050, engine='openpyxl') as writer:
        
        for scenario in scenario_order:
            if scenario not in scenarios:
                continue
                
            sheet_name = scenario_names.get(scenario, scenario[:10])
            print(f"Processing scenario: {scenario} (Sheet: {sheet_name}) for 2030-2050")
            
            # Extract capacity data for this scenario
            scenario_cap_data = filtered_cap_data[filtered_cap_data['Scenario'] == scenario]
            
            # Extract special technology data
            biomass_scenario = biomass_cap_data[biomass_cap_data['Scenario'] == scenario]
            bio_ccs_scenario = bio_ccs_cap_data[bio_ccs_cap_data['Scenario'] == scenario]
            gas_ccs_scenario = gas_ccs_cap_data[gas_ccs_cap_data['Scenario'] == scenario]
            coal_ccs_scenario = coal_ccs_cap_data[coal_ccs_cap_data['Scenario'] == scenario]
            
            # Skip if no data for this scenario
            if len(scenario_cap_data) == 0:
                print(f"No data for scenario: {scenario}, skipping...")
                continue
            
            # Calculate new installed capacity for 2030-2050
            new_installed_cap = calculate_new_installed_capacity(
                scenario_cap_data, biomass_scenario, bio_ccs_scenario, 
                gas_ccs_scenario, coal_ccs_scenario, 2030, 2050
            )
            
            # Store the data for the matrix DataFrame
            matrix_dfs_2050[scenario] = new_installed_cap.copy()
            
            # Process hydro generation data for 2050
            hydro_gen_2050 = filtered_gen_data[(filtered_gen_data['Scenario'] == scenario) & 
                                             (filtered_gen_data['Tech'] == 'Hydro') & 
                                             (filtered_gen_data['Period'] == 2050)]
            hydro_act_series = hydro_gen_2050.groupby('RegionClass')['Pv'].sum()
            hydro_act = pd.Series(0, index=new_installed_cap.index)
            for region in hydro_act_series.index:
                if region in hydro_act.index:
                    hydro_act[region] = hydro_act_series[region]
            
            # Calculate additional columns
            new_installed_cap['Sum'] = new_installed_cap.drop(columns=['Storage'], errors='ignore').sum(axis=1)
            
            with np.errstate(divide='ignore', invalid='ignore'):
                new_installed_cap['Storage_Ratio'] = new_installed_cap['Storage'] / new_installed_cap['Sum']
            
            new_installed_cap['Wind_Solar_Sum'] = new_installed_cap['Wind'] + new_installed_cap['Solar']
            
            with np.errstate(divide='ignore', invalid='ignore'):
                new_installed_cap['Wind_Solar_Ratio'] = new_installed_cap['Wind_Solar_Sum'] / new_installed_cap['Sum']
            
            new_installed_cap = new_installed_cap.fillna(0)
            new_installed_cap['Hydro_act'] = hydro_act
            
            # Calculate total row
            total_series = {}
            for col in new_installed_cap.columns:
                if col not in ['Storage_Ratio', 'Wind_Solar_Ratio']:
                    total_series[col] = new_installed_cap[col].sum()
            
            with np.errstate(divide='ignore', invalid='ignore'):
                total_series['Storage_Ratio'] = total_series.get('Storage', 0) / total_series.get('Sum', 1) if total_series.get('Sum', 0) != 0 else 0
                total_series['Wind_Solar_Ratio'] = total_series.get('Wind_Solar_Sum', 0) / total_series.get('Sum', 1) if total_series.get('Sum', 0) != 0 else 0
            
            # Convert to DataFrame for Excel output
            new_installed_cap_df = new_installed_cap.reset_index()
            new_installed_cap_df.rename(columns={'index': 'RegionClass'}, inplace=True)
            
            # Append total row
            total_df = pd.DataFrame([['Total'] + [total_series.get(col, 0) for col in new_installed_cap.columns]], 
                                   columns=['RegionClass'] + list(new_installed_cap.columns))
            new_installed_cap_df = pd.concat([new_installed_cap_df, total_df], ignore_index=True)
            new_installed_cap_df = new_installed_cap_df.fillna(0)
            
            # Write to Excel
            new_installed_cap_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the worksheet
            format_worksheet(writer.sheets[sheet_name], new_installed_cap_df)
    
    print(f"New installed capacity Excel files created successfully:")
    print(f"1. 2020-2030 period: {output_path_2030}")
    print(f"2. 2030-2050 period: {output_path_2050}")
    
    return output_path_2030, output_path_2050, matrix_dfs_2030, matrix_dfs_2050

def format_worksheet(worksheet, df):
    """Format worksheet with proper styling"""
    # Auto-adjust column widths
    for idx, col in enumerate(df.columns):
        column_letter = get_column_letter(idx + 1)
        column_width = max(len(str(col)), df[col].astype(str).str.len().max())
        worksheet.column_dimensions[column_letter].width = column_width + 2
        
    # Apply number formatting to numeric columns
    for row in range(2, len(df) + 2):  # +2 for header and 1-indexing
        for col in range(2, len(df.columns) + 1):  # +1 for 1-indexing
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                # Format ratios as percentages
                if df.columns[col-1] in ['Storage_Ratio', 'Wind_Solar_Ratio']:
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '#,##0.00'
    
    # Add a light header formatting
    header_row = worksheet[1]
    for cell in header_row:
        cell.font = Font(bold=True)

if __name__ == "__main__":
    output_2030, output_2050, matrix_2030, matrix_2050 = generate_tiam_new_installed_excel()